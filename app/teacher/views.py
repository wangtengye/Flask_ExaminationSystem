import os

import itertools
from flask import Blueprint, render_template, redirect, request, send_from_directory, session, json
from flask_login import login_required, current_user
from openpyxl import load_workbook
from pyecharts import Line
from werkzeug.utils import secure_filename

from .. import db

from ..models import Choice, Page, table_dict, Judge, Subject, Class, Paper, Student, Record

teacher = Blueprint('teacher', __name__)


@teacher.before_request
@login_required
def check_permission():
    if current_user.permission != 'Teacher':
        return redirect('/not_permission')


@teacher.route('/')
def home():
    return render_template('teacher/teacher_page.html')


@teacher.route('/question_list')
def question_list():
    choice, p = get_question('choice')
    return render_template('teacher/question_list.html', choice=choice, page=p)


@teacher.route('/truefalse_list')
def truefalse_list():
    choice, p = get_question('judge')
    return render_template('teacher/truefalse_list.html', choice=choice, page=p)


@teacher.route('/sub_list')
def sub_list():
    choice, p = get_question('sub')
    return render_template('teacher/sub_list.html', choice=choice, page=p)


def get_question(qtype):
    current_page = request.args.get('currentPage') if request.args.get('currentPage') is not None else 1
    count = table_dict[qtype].query.count()

    p = Page()
    total_page = count // p.page_number if count % p.page_number == 0 else count // p.page_number + 1

    p.count = count
    p.total_page = total_page
    p.current_page = current_page

    que = table_dict[qtype].query.offset((int(current_page) - 1) * p.page_number).limit(p.page_number).all()
    return que, p


@teacher.route('/<qtype>/delete')
def delete(qtype):
    que = table_dict[qtype].query.get(request.args['id'])
    db.session.delete(que)
    db.session.commit()
    return 'true'


@teacher.route('/<qtype>/delete_batch', methods=['post'])
def delete_batch(qtype):
    ids = request.values.getlist('list[]')
    table_dict[qtype].query.filter(table_dict[qtype].id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return 'true'


@teacher.route('/<qtype>/download')
def download(qtype):
    filename = qtype + '_template.xlsx'
    return send_from_directory('static/excel', filename, as_attachment=True)


@teacher.route('/<qtype>/upload', methods=['post'])
def upload(qtype):
    file = request.files['file']
    filename = secure_filename(file.filename)
    path = os.path.join(os.path.dirname(__file__), filename)
    file.save(path)

    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if qtype == 'choice':
            que = Choice(question=row[0].value, option1=row[1].value, option2=row[2].value,
                         option3=row[3].value, option4=row[4].value, rightanswer=row[5].value)
        elif qtype == 'judge':
            que = Judge(question=row[0].value, rightanswer=row[1].value)
        elif qtype == 'sub':
            que = Subject(question=row[0].value, refanswer=row[1].value)
        db.session.add(que)
    db.session.commit()
    os.remove(path)
    return "true"


@teacher.route('/<qtype>/update', methods=['post'])
def update(qtype):
    data = request.form
    que = table_dict[qtype].query.get(data['id'])
    que.question = data['question']
    if qtype == 'choice':
        que.option1 = data['option1']
        que.option2 = data['option2']
        que.option3 = data['option3']
        que.option4 = data['option4']
        que.rightanswer = data['rightAnswer']
    elif qtype == 'judge':
        que.rightanswer = data['rightAnswer']
    elif qtype == 'sub':
        que.refanswer = data['refAnswer']
    db.session.add(que)
    db.session.commit()
    return 'true'


@teacher.route('/get_paper')
def get_paper():
    return render_template('teacher/choice_combine.html', choice=Choice.query.all())


@teacher.route('/<qtype>/combine')
def combine(qtype):
    if qtype == 'choice':
        session['choi'] = request.args.get('choice')
        return render_template('teacher/judge_combine.html', judge=Judge.query.all())
    elif qtype == 'judge':
        session['judg'] = request.args['judge']
        return render_template('teacher/sub_combine.html', sub=Subject.query.all())
    elif qtype == 'sub':
        session['sub'] = request.args['sub']
        return render_template('teacher/other_config.html')


@teacher.route('/paper_finish', methods=['post'])
def paper_finish():
    name = request.args['name']
    begintime = request.args['begintime']
    finishtime = request.args['finishtime']
    choi = session['choi']
    judg = session['judg']
    sub = session['sub']
    tid = current_user.id
    classid = '[|' + str(Class.get_id_by_name(request.args['classid'])) + '|]'

    choicescore = '4,' * len(choi.split(','))
    judgscore = '4,' * len(judg.split(','))
    subscore = '4,' * len(sub.split(','))

    print(name, begintime, finishtime, choi, judg, sub, tid, classid, choicescore, judgscore, subscore)
    paper = Paper(name, begintime, finishtime, choi, judg, sub, choicescore[:-1],
                  judgscore[:-1], subscore[:-1], tid, classid)
    db.session.add(paper)
    db.session.commit()
    return 'true'


@teacher.route('/student_grade')
def student_grade():
    paper = Paper.query.filter_by(tid=current_user.id)
    return render_template('teacher/student_grade.html', paper=paper)


@teacher.route('/student_grade_list')
def student_grade_list():
    grade = db.session.query(Student.name, Record.score).filter(Record.pid == request.args['pid']) \
        .filter(Student.id == Record.sid).all()
    return render_template('teacher/student_grade_list.html', name=request.args['name'], grade=grade)


@teacher.route('/get_chart')
def get_chart():
    raw_data = db.session.query(Record.score).filter(Record.pid == request.args['pid']).all()
    data = list(itertools.chain(*raw_data))
    line = Line(request.args['name'])
    line.add(None, is_xaxis_show=False, x_axis=[''] * len(data), is_xaxis_boundarygap=False, y_axis=data,
             mark_point=['min', 'max'], mark_line=['average'])
    return render_template('teacher/student_chart.html', name=request.args['name'],
                           data=json.dumps(line.options))
