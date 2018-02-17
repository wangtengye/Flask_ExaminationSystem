import os
from flask import Blueprint, render_template, redirect, request, send_from_directory
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from .. import db

from ..models import Student, Teacher, Class, table_dict
from openpyxl import load_workbook

admin = Blueprint('admin', __name__)


@admin.before_request
@login_required
def check_permission():
    if current_user.permission != 'Admin':
        return redirect('/not_permission')


@admin.route('/')
def home():
    return render_template('admin/admin_page.html')


@admin.route('/add_student', methods=['get', 'post'])
def add_student():
    if request.is_xhr:
        stu = Student(account=request.form['account'], name=request.form['name']
                      , class_id=Class.get_id_by_name(request.form['classid']), password='123456')
        db.session.add(stu)
        db.session.commit()
        return 'true'
    return render_template('admin/add_student.html')


@admin.route('/add_teacher', methods=['get', 'post'])
def add_teacher():
    if request.is_xhr:
        tea = Teacher(account=request.form['account'], name=request.form['name'], password='123456')
        db.session.add(tea)
        db.session.commit()
        return 'true'
    return render_template('admin/add_teacher.html')


@admin.route('/<usr>/download')
def download(usr):
    filename = usr + '_template.xlsx'
    return send_from_directory('static/excel', filename, as_attachment=True)


@admin.route('/<usr>/upload', methods=['post'])
def upload(usr):
    file = request.files['file']
    filename = secure_filename(file.filename)
    path = os.path.join(os.path.dirname(__file__), filename)
    file.save(path)

    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if usr == 'student':
            stu = Student(account=row[0].value, name=row[1].value,
                          class_id=Class.get_id_by_name(row[2].value), password="123456")
        elif usr == 'teacher':
            stu = Teacher(account=row[0].value, name=row[1].value, password="123456")
    db.session.add(stu)
    db.session.commit()
    os.remove(path)
    return "true"


@admin.route('/student_manage')
def student_manage():
    students = Student.query.all()
    for stu in students:
        stu.class_name = Class.get_name_by_id(stu.class_id)
    return render_template('admin/student_manage.html', students=students)


@admin.route('/teacher_manage')
def teacher_manage():
    teachers = Teacher.query.all()
    return render_template('admin/teacher_manage.html', teachers=teachers)


@admin.route('/<usr>/reset', methods=['post'])
def reset_password(usr):
    usr = table_dict[usr].query.get(request.args['id'])
    usr.password = '123456'
    db.session.add(usr)
    db.session.commit()
    return 'true'


@admin.route('/<usr>/delete')
def delete(usr):
    usr = table_dict[usr].query.get(request.args['id'])
    db.session.delete(usr)
    db.session.commit()
    return 'true'


@admin.route('/<usr>/delete_batch', methods=['post'])
def delete_batch(usr):
    ids = request.values.getlist('list[]')
    table_dict[usr].query.filter(table_dict[usr].id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return 'true'


@admin.route('/<usr>/update', methods=['post'])
def update(usr):
    data = request.form
    user = table_dict[usr].query.get(data['id'])
    if usr == 'student':
        user.class_id = Class.get_id_by_name(data['classid'])
    user.account = data['account']
    user.name = data['name']
    db.session.add(user)
    db.session.commit()
    return 'true'


@admin.route('/admin_info')
def admin_info():
    return render_template('admin/admin_info.html')
